# Пропишите функции для ботов здесь
import base64
import io
import json
import os
import uuid

import httplib2
import openpyxl
import requests
from apps.bot.models import TelegramUser
from apps.bot.utils.constants import SPREADSHEET_ID
from django.conf import settings
from googleapiclient.discovery import build
from loguru import logger
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from PIL import Image

logger.add('logs/bot_tools.log')


class WBPersonalApiClient:
    def __init__(self, supplierId: str = None, WBToken: str = None) -> None:
        self.supplierId = supplierId
        self.WBToken = WBToken
        self.base_url = 'https://seller.wildberries.ru/'

    def _headers(self):
        return {
            'Accept': '*/*',
            'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,bg;q=0.6',
            'Connection': 'keep-alive',
            'Content-type': 'application/json',
            'Cookie': f'x-supplier-id={self.supplierId};WBToken={self.WBToken};',
            'Host': 'seller.wildberries.ru',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Safari/605.1.15'
        }

    def make_request(self, method: str, url: str, params: tuple = None, payload: dict = None):
        try:
            return requests.request(method, url=url, params=params, headers=self._headers(), data=json.dumps(payload) if payload is not None else None)
        except Exception as err:
            logger.error(err, exc_info=True)
            return False

    def _check_api(self):
        pass

    def send_login_code(self, phone: str):
        url = self.base_url + 'passport/api/v2/auth/login_by_phone'

        payload = {
            'phone': str(phone).replace('+', ''),
            'is_terms_and_conditions_accepted': True
        }

        response = self.make_request(
            'POST', url=url, payload=payload
        )
        if response is False:
            return False, 'Ошибка подключения'

        if response.status_code == 200:
            logger.success('Success send verify code to %s' % phone)
            return True, response.json()
        logger.error('Error sending verify code %i\n[RESPONSE TEXT]: %s\n[RESPONSE PAYLOAD]: %s' % (response.status_code, response.text, payload))
        return False, 'Неверный номер телефона'

    def verify_login_code(self, token: str, code: str):
        url = self.base_url + 'passport/api/v2/auth/login'

        payload = {
            'options': {
                'notify_code': str(code)
            },
            'token': token,
            'device': 'Macintosh,Google Chrome 110.0'
        }

        response = self.make_request(
            'POST', url=url, payload=payload
        )
        if response is False:
            return False, 'Ошибка подключения'

        if response.status_code == 200:
            logger.success('Success verify code (TOKEN: %s)' % response.cookies['WBToken'])
            return True, {
                'WBToken': response.cookies['WBToken']
            }
        if response.json().get('error', '') == 'invalid_token':
            return False, 'Неверный токен'
        elif response.json().get('error', '') == 'invalid_code':
            return False, 'Неверный код подтверждения'
        logger.error('Error verify code %s' % response.json().get('error', ''))
        return False, response.json().get('error', '')

    def get_suppliers(self):
        url = self.base_url + 'ns/suppliers/suppliers-portal-core/suppliers'

        payload = [
            {
                'method': 'getUserSuppliers',
                'params': {},
                'id': 'json-rpc_4',
                'jsonrpc': '2.0'
            }
        ]
        response = self.make_request(
            'POST', url=url, payload=payload
        )
        if response is False:
            return False, 'Ошибка подключения'

        if response.status_code == 200:
            logger.success('Success get suppliers by token, total suppliers: %i' % len(response.json()[0]['result']['suppliers']))
            return True, response.json()[0]['result']['suppliers']
        logger.error('Error getting suppliers: %s' % response.text)
        return False, 'Не удалось получить продавцов кабинета Wildberries'

    def get_feedbacks(self, skip: int = 0, take: int = 50):
        url = self.base_url + 'ns/api/suppliers-portal-feedbacks-questions/api/v1/feedbacks'
        params = (
            ('isAnswered', False, ),
            ('metaDataKeyMustNot', 'norating', ),
            ('order', 'dateDesc', ),
            ('skip', skip, ),
            ('take', take, ),
        )

        response = self.make_request(
            'GET', url=url, params=params
        )
        if response is False:
            return False, 'Ошибка подключения'

        if response.status_code == 200:
            logger.success('Success getting feedbacks, total feedbacks: %i' % len(response.json().get('data', {}).get('feedbacks', [])))
            return True, response.json().get('data', {}).get('feedbacks', [])
        logger.error('Invalid key get feedbacks %s' % response.text)
        return False, 'Неверный ключ'

    def get_cards(self, search: str = ''):
        url = self.base_url + 'ns/viewer/content-card/viewer/tableList'

        offset = 100

        def get_page(skip=0):
            payload = {
                'sort': {
                    'limit': offset,
                    'offset': skip,
                    'searchValue': search,
                    'sortColumn': 'updateAt',
                    'ascending': False
                },
                'filter': {
                    'tags': [],
                    'brands': [],
                    'subjects': [],
                    'hasPhoto': 0
                }
            }
            response = self.make_request(
                'POST', url=url, payload=payload
            )
            return response

        response = get_page()
        if response is False:
            return False, 'Ошибка подключения'

        cards = []
        batch = response.json()
        if len(batch.get('data', {}).get('cards', [])) == 0:
            return False, 'Артикул не найден'
        cards += batch.get('data', {}).get('cards', [])

        attempt = 1
        while True:
            response = get_page(offset*attempt)
            attempt += 1
            if response is False:
                return False, 'Ошибка подключения'
            if response.status_code == 200:
                logger.success('Success getting cards, total cards: %i, attempt: %i' % (len(cards), attempt))
                if len(response.json().get('data', {}).get('cards', [])) == 0:
                    break
                cards += response.json().get('data', {}).get('cards')
            else:
                logger.error('Invalid key get cards %s' % response.status_code)
                return False, 'Неверный ключ'
        return True, cards


def get_suppliers(user: TelegramUser):
    if user.WBToken is None:
        return None, 'Вы не авторизованы в кабинете Wildberries'
    client = WBPersonalApiClient(WBToken=user.WBToken)
    response = client.get_suppliers()
    if response[0]:
        for supplier in response[1]:
            if not user.personal_set.filter(supplierId=supplier['id']).exists():
                personal = user.personal_set.create(
                    user=user,
                    supplierId=supplier['id'],
                    oldId=supplier['oldID'],
                    name=supplier['name'],
                    full_name=supplier['fullName']
                )
                logger.success(f'New personal added: {personal} {personal.supplierId}')
        return True, user.personal_set.all()
    else:
        return False, response[1]


def get_tracked_articles(user: TelegramUser):
    if len(user.personal_set.all()) != 0:
        total_tracked_articles = []
        for personal in user.personal_set.all():
            total_tracked_articles += [article for article in personal.get_tracked_articles()]
        if len(total_tracked_articles) == 0:
            return None, 'У вас нет отслеживаемых артикулов'
        logger.success('Success getting tracked articles, total tracked articles: %i' % len(total_tracked_articles))
        return True, total_tracked_articles
    else:
        return False, 'У вас нет добавленых кабинетов Wildberries'


def scale_sheet(sheet: object, plus_size: int = 10):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + plus_size))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value


def get_tracked_articles_excel(user: TelegramUser, to_delete: bool = False):
    tracked_articles = get_tracked_articles(user)
    if tracked_articles[0] is None or tracked_articles[0] is False:
        return tracked_articles
    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        'Артикул WB', 'Артикул продавца'
    ]

    if to_delete:
        headers.append('Заполните да, если товар НЕ нужно отслеживать')

    sheet.cell(row=1, column=1).value = 'Список отслеживаемых артикулов'
    for i, header in enumerate(headers):
        sheet.cell(row=2, column=i+1).value = header
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(headers))
    for row in range(1, 3):
        for column in range(1, len(headers)+1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill(start_color=Color('B1A0C7'), end_color=Color('B1A0C7'), fill_type='solid')
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    row = 3
    for article in tracked_articles[1]:
        sheet.cell(row=row, column=1).value = article.nmId
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=1).hyperlink = f'https://www.wildberries.ru/catalog/{article.nmId}/detail.aspx?targetUrl=SP'
        sheet.cell(row=row, column=2).value = article.article
        row += 1
    scale_sheet(sheet)
    sheet.column_dimensions['A'].width = 15
    if to_delete:
        sheet.column_dimensions['C'].width = 28
    file_stream = io.BytesIO()
    wb.save(file_stream)

    file_stream.seek(0)
    return True, base64.b64encode(file_stream.read()).decode('utf-8')


def get_stars_display(user: TelegramUser):
    return '⭐️' * user.notification_stars


def get_personals_pages_current_page(keyboard: object):
    return int(keyboard[-2][0].callback_data.split(':')[-1])


def get_personal_cards_excel(user: TelegramUser, personal: object):
    if user.WBToken is None:
        return False, 'Вы не авторизованы в кабинете Wildberries'
    client = WBPersonalApiClient(personal.supplierId, user.WBToken)
    cards = client.get_cards()
    if cards[0] is False:
        return cards
    if len(cards[1]) == 0:
        return None, 'В данном кабинете нет карточек'
    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        'Артикул WB', 'Артикул продавца', 'Бренд', 'Предмет', 'Цвет', 'Размер', 'Баркод', 'Заполните да, если нужно отслеживать товар'
    ]

    sheet.cell(row=1, column=1).value = 'Список товаров %s' % personal.name
    for i, header in enumerate(headers):
        sheet.cell(row=2, column=i+1).value = header
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(headers))
    for row in range(1, 3):
        for column in range(1, len(headers)+1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill(start_color=Color('B1A0C7'), end_color=Color('B1A0C7'), fill_type='solid')
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    row = 3
    for card in cards[1]:
        sheet.cell(row=row, column=1).value = card['nmID']
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=1).hyperlink = f'https://www.wildberries.ru/catalog/{card["nmID"]}/detail.aspx?targetUrl=SP'
        sheet.cell(row=row, column=2).value = card['vendorCode']
        sheet.cell(row=row, column=3).value = card['Бренд']
        sheet.cell(row=row, column=4).value = card['Предмет']
        sheet.cell(row=row, column=5).value = ', '.join(card['Цвет']) if len(card['Цвет']) != 0 else '-'
        sheet.cell(row=row, column=6).value = card['size'][0].get('wbSize') if card['size'][0].get('wbSize') != '' else '-'
        sheet.cell(row=row, column=7).value = card['size'][0].get('skus', [''])[0]
        if personal.trackedarticle_set.filter(nmId=card['nmID']).exists():
            sheet.cell(row=row, column=8).value = 'Уже отслеживается'

        row += 1
    scale_sheet(sheet)

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 28

    file_stream = io.BytesIO()
    wb.save(file_stream)

    file_stream.seek(0)
    return True, base64.b64encode(file_stream.read()).decode('utf-8')


def add_articles_to_track(personal: object, file: str):
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return False, 'Это не похоже на файл Excel'
    sheet = wb.active
    added_articles = []
    for row in range(3, sheet.max_row+1):
        if sheet.cell(row=row, column=8).value is not None and sheet.cell(row=row, column=8).value.lower() == 'да':
            added_articles.append(
                personal.trackedarticle_set.create(
                    nmId=sheet.cell(row=row, column=1).value,
                    article=sheet.cell(row=row, column=2).value,
                    brand=sheet.cell(row=row, column=3).value
                )
            )
    return True, added_articles


def remove_articles_from_track(user: TelegramUser, file: str):
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return False, 'Это не похоже на файл Excel'
    sheet = wb.active
    removed_articles = []
    for row in range(2, sheet.max_row+1):
        nmId = sheet.cell(row=row, column=1).value
        if sheet.cell(row=row, column=3).value is not None and sheet.cell(row=row, column=3).value.lower() == 'да':
            for personal in user.personal_set.all():
                if personal.trackedarticle_set.filter(nmId=str(nmId)).exists():
                    tracked_article = personal.trackedarticle_set.get(nmId=str(nmId))
                    removed_articles.append(tracked_article)
                    tracked_article.delete()
    return True, removed_articles


def delete_files(path: list):
    for file in path:
        os.remove(file)


def merge_card_images(photos: list):
    images_paths = []
    for photo in photos:
        r = requests.get(photo)
        im_fname = os.path.join(settings.STATICFILES_DIRS[0], '{}.{}'.format(uuid.uuid4().hex, photo.split('.')[-1]))
        with open(im_fname, 'wb') as f:
            f.write(r.content)
        images_paths.append(im_fname)

    images = [Image.open(x) for x in images_paths]

    widths, heights = zip(*(i.size for i in images))

    total_width = sum(widths)
    max_height = max(heights)

    new_im = Image.new('RGB', (total_width, max_height))

    x_offset = 0
    for im in images:
        new_im.paste(im, (x_offset, 0))
        x_offset += im.size[0]
    file_stream = io.BytesIO()
    new_im.save(file_stream, 'png')
    file_stream.seek(0)
    delete_files(images_paths)
    return base64.b64encode(file_stream.read()).decode('utf-8')


def get_service_sacc():
    creds_json = os.path.dirname(__file__) + '/creds/push-feedback-wb-bot-6a2cdc52ebc8.json'
    scopes = ['https://www.googleapis.com/auth/spreadsheets']

    creds_service = ServiceAccountCredentials.from_json_keyfile_name(creds_json, scopes).authorize(httplib2.Http())
    return build('sheets', 'v4', http=creds_service)


def reset_sheet(service, sheet_id):
    body = {
        'requests': [
            {
                'updateCells': {
                    'range': {
                        'sheetId': sheet_id
                    },
                    'fields': 'userEnteredValue'
                }
            }
        ]
    }
    return service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()


def delete_sheet(service, sheet_id):
    body = {
        'requests': [
            {
                'deleteSheet': {
                    'sheetId': sheet_id
                }
            }
        ]
    }

    return service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()


def delete_sheets(service=None, requests=[]):
    if service is None:
        service = get_service_sacc()
    body = {
        'requests': requests
    }
    return service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()


def add_sheet(service=None, title='Новый лист'):
    if service is None:
        service = get_service_sacc()
    body = {
        'requests': [
            {
                'addSheet': {
                    'properties': {
                        'title': title,
                        'gridProperties': {
                            'rowCount': 999999,
                            'columnCount': 4
                        }
                    }
                }
            }
        ]
    }

    return service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()


def get_spreadsheets(service=None):
    if service is None:
        service = get_service_sacc()
    sheet_metadata = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets = sheet_metadata.get('sheets', '')
    print(sheets)
    return sheets
    # title = sheets[0].get("properties", {}).get("title", "Sheet1")
    # sheet_id = sheets[0].get("properties", {}).get("sheetId", 0)


def append_table_values(service=None, sheet_title='', values=[]):
    if service is None:
        service = get_service_sacc()
    body = {
        'values': values
    }
    result = service.spreadsheets().values().append(spreadsheetId=SPREADSHEET_ID, range=f'{sheet_title}!A:C', valueInputOption='USER_ENTERED', body=body).execute()
    print(f"{(result.get('updates').get('updatedCells'))} cells appended.")
    return result


def auto_resize_sheet(service=None, sheetId=''):
    body = {
        'requests': [
            {
                'autoResizeDimensions': {
                    'dimensions': {
                        'sheetId': sheetId,
                        'dimension': 'COLUMNS',
                        'startIndex': 0,
                        'endIndex': 2
                    }
                }
            }
        ]
    }
    return service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
